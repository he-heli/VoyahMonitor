from __future__ import annotations

import tempfile
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from voyah_monitor.storage import DailyMileage, SnapshotRecord, TelemetryStorage
from voyah_monitor.timeutil import format_moscow, moscow_date, moscow_now
from voyah_monitor.vehicle_status import SNAPSHOT_EXPORT_DATA_HEADERS, snapshot_export_field_map


def _export_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    return value


def _snapshot_row(record: SnapshotRecord) -> list[Any]:
    telemetry = record.telemetry
    fields = snapshot_export_field_map(telemetry)
    return [
        record.id,
        format_moscow(telemetry.captured_at),
        *[_export_cell(fields.get(header)) for header in SNAPSHOT_EXPORT_DATA_HEADERS],
    ]


SNAPSHOT_HEADERS = ("ID", "Время") + SNAPSHOT_EXPORT_DATA_HEADERS

MILEAGE_HEADERS = (
    "День",
    "Автомобиль",
    "Пробег начало, км",
    "Пробег конец, км",
    "Пробег за день, км",
)


def _mileage_row(row: DailyMileage) -> list[Any]:
    return [
        row.day.isoformat(),
        row.vehicle_key,
        row.start_odometer_km,
        row.end_odometer_km,
        row.distance_km,
    ]


def _snapshot_source(
    storage: TelemetryStorage,
    *,
    days: int | None,
    snapshots: list[SnapshotRecord] | None,
) -> Iterator[SnapshotRecord]:
    if snapshots is not None:
        yield from snapshots
        return
    yield from storage.iter_snapshots_in_range(days=days)


def export_history_xlsx_to_path(
    storage: TelemetryStorage,
    dest: Path,
    *,
    days: int | None = None,
    snapshots: list[SnapshotRecord] | None = None,
) -> int:
    """Stream rows to disk; returns snapshot count written."""
    mileage_rows = _filter_mileage(storage.all_daily_mileage(), days=days)
    snapshot_iter = _snapshot_source(storage, days=days, snapshots=snapshots)

    workbook = Workbook(write_only=True)
    snapshots_sheet = workbook.create_sheet("Снимки")
    snapshots_sheet.append(SNAPSHOT_HEADERS)
    count = 0
    for record in snapshot_iter:
        snapshots_sheet.append(_snapshot_row(record))
        count += 1

    mileage_sheet = workbook.create_sheet("Пробег по дням")
    mileage_sheet.append(MILEAGE_HEADERS)
    for row in mileage_rows:
        mileage_sheet.append(_mileage_row(row))

    dest.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dest)
    return count


def export_history_xlsx(
    storage: TelemetryStorage,
    *,
    days: int | None = None,
    snapshots: list[SnapshotRecord] | None = None,
    dest: Path | None = None,
) -> bytes:
    if dest is not None:
        export_history_xlsx_to_path(storage, dest, days=days, snapshots=snapshots)
        return dest.read_bytes()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
    try:
        export_history_xlsx_to_path(storage, path, days=days, snapshots=snapshots)
        return path.read_bytes()
    finally:
        path.unlink(missing_ok=True)


def export_filename(*, days: int | None = None) -> str:
    stamp = moscow_now().strftime("%Y%m%d_%H%M")
    if days is None:
        return f"voyah_history_all_{stamp}.xlsx"
    return f"voyah_history_{days}d_{stamp}.xlsx"


def period_label(days: int | None) -> str:
    if days is None:
        return "за всё время"
    if days == 1:
        return "за 1 день"
    if days % 30 == 0 and days >= 30:
        months = days // 30
        return f"за {months} мес."
    return f"за {days} дн."


def _filter_mileage(rows: list[DailyMileage], *, days: int | None) -> list[DailyMileage]:
    if days is None:
        return rows
    from datetime import timedelta

    min_day = moscow_date() - timedelta(days=days)
    return [row for row in rows if row.day >= min_day]
