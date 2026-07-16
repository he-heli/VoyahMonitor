from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from voyah_monitor.history_export import SNAPSHOT_HEADERS, export_history_xlsx
from voyah_monitor.storage import TelemetryStorage
from voyah_monitor.telemetry import VehicleTelemetry, format_status
from voyah_monitor.timeutil import format_moscow, to_moscow


def test_format_moscow_from_utc() -> None:
    dt = datetime(2026, 7, 16, 20, 41, 41, tzinfo=UTC)
    assert format_moscow(dt) == "2026-07-16 23:41:41"
    assert to_moscow(dt).tzname() in {"MSK", "Europe/Moscow"}


def test_export_and_status_use_moscow_time(tmp_path) -> None:
    storage = TelemetryStorage(tmp_path / "test.db")
    captured = datetime(2026, 7, 16, 20, 41, 41, tzinfo=UTC)
    storage.save_snapshot(
        VehicleTelemetry(
            captured_at=captured,
            vehicle_id="car1",
            battery_percent=70.0,
        )
    )

    text = format_status(
        VehicleTelemetry(captured_at=captured, vehicle_id="car1", battery_percent=70.0)
    )
    assert "2026-07-16 23:41:41" in text

    workbook = load_workbook(BytesIO(export_history_xlsx(storage)))
    sheet = workbook["Снимки"]
    time_col = SNAPSHOT_HEADERS.index("Время") + 1
    assert sheet.cell(row=2, column=time_col).value == "2026-07-16 23:41:41"
