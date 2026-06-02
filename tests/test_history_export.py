from __future__ import annotations

from datetime import UTC, datetime, timedelta

from openpyxl import load_workbook

from voyah_monitor.history_export import SNAPSHOT_HEADERS, export_history_xlsx
from voyah_monitor.storage import TelemetryStorage
from voyah_monitor.telemetry import VehicleTelemetry
from voyah_monitor.vehicle_status import SNAPSHOT_EXPORT_DATA_HEADERS


def _dashboard_raw(
    *,
    fuel_pct: float = 45.0,
    fuel_km: float = 320.0,
) -> dict:
    return {
        "table": {
            "vin": "VIN1",
            "licensePlate": "А123АА777",
            "carModel": {"displayName": "Test Car", "modname": "EV"},
            "liveSensors": {
                "fuelPercentage": fuel_pct,
                "remainsMileageFuel": fuel_km,
                "batteryPercentage": 70.0,
                "remainsMileage": 280.0,
                "odometer": 1000.0,
                "soh": 100.0,
                "12VBatteryVoltage": 12.8,
                "coolantTemp": 22.0,
                "batteryTemp": 18.0,
                "outsideTemp": 5.0,
            },
        },
        "geo": {},
        "detail": {},
        "tbox": {
            "isOnline": True,
            "sensors": {"chip": {"title": "Доступен"}},
        },
    }


def test_export_history_xlsx_contains_snapshots(tmp_path) -> None:
    db_path = tmp_path / "test.db"
    storage = TelemetryStorage(db_path)
    now = datetime.now(UTC)
    storage.save_snapshot(
        VehicleTelemetry(
            captured_at=now - timedelta(days=1),
            vehicle_id="car1",
            vin="VIN1",
            name="Test Car",
            odometer_km=1000.0,
            battery_percent=70.0,
            v12_voltage=12.8,
            soh_percent=100.0,
            is_online=True,
            status="Доступен",
            raw=_dashboard_raw(),
        )
    )

    data = export_history_xlsx(storage, days=None)
    assert data.startswith(b"PK")

    from io import BytesIO

    workbook = load_workbook(BytesIO(data))
    assert "Снимки" in workbook.sheetnames
    sheet = workbook["Снимки"]
    assert sheet.max_row >= 2
    assert tuple(cell.value for cell in sheet[1]) == SNAPSHOT_HEADERS
    assert "Топливо, %" in SNAPSHOT_EXPORT_DATA_HEADERS
    assert "Топливо, км" in SNAPSHOT_EXPORT_DATA_HEADERS

    fuel_pct_col = SNAPSHOT_HEADERS.index("Топливо, %") + 1
    fuel_km_col = SNAPSHOT_HEADERS.index("Топливо, км") + 1
    assert sheet.cell(row=2, column=fuel_pct_col).value == 45.0
    assert sheet.cell(row=2, column=fuel_km_col).value == 320.0


def test_export_fallback_without_dashboard_raw(tmp_path) -> None:
    db_path = tmp_path / "test.db"
    storage = TelemetryStorage(db_path)
    storage.save_snapshot(
        VehicleTelemetry(
            vehicle_id="car1",
            battery_percent=80.0,
            range_km=300.0,
        )
    )

    from io import BytesIO

    workbook = load_workbook(BytesIO(export_history_xlsx(storage)))
    sheet = workbook["Снимки"]
    battery_col = SNAPSHOT_HEADERS.index("Батарея, %") + 1
    assert sheet.cell(row=2, column=battery_col).value == 80.0
